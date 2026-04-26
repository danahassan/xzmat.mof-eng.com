import csv
import io
from datetime import datetime, timedelta
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, abort, Response)
from flask_login import login_required, current_user
from models import db, User, Message, Application, ROLE_ADMIN, ROLE_SUPERVISOR, ROLE_USER, ROLE_EMPLOYER
from helpers import push_notification
from sqlalchemy import or_, and_

messages_bp = Blueprint('messages', __name__)


def _is_coordinator(user):
    return user.role in (ROLE_SUPERVISOR, 'university_coordinator')


def _is_student(user):
    return user.role in (ROLE_USER, 'student')


def _is_company_user(user):
    return user.role == ROLE_EMPLOYER


def _recipient_matches_channel(receiver, channel):
    if channel == 'company_coordinator':
        return _is_company_user(receiver) or _is_coordinator(receiver)
    if channel == 'coordinator_student':
        return _is_coordinator(receiver) or _is_student(receiver)
    return True


def _can_message(sender, receiver):
    """Check if sender is allowed to message receiver based on roles."""
    if sender.id == receiver.id:
        return False
    if not receiver.is_active:
        return False
    # Admin can message anyone
    if sender.role == ROLE_ADMIN:
        return True
    # Anyone can message admin
    if receiver.role == ROLE_ADMIN:
        return True
    # Coordinator can message coordinator
    if _is_coordinator(sender) and _is_coordinator(receiver):
        return True
    # Company can message coordinator, and coordinator can message company
    if (_is_company_user(sender) and _is_coordinator(receiver)) or (_is_coordinator(sender) and _is_company_user(receiver)):
        return True
    # Coordinator and student chat channel
    if (_is_coordinator(sender) and _is_student(receiver)) or (_is_student(sender) and _is_coordinator(receiver)):
        return True
    # Backward-compatible assignment rule for old data flows
    if sender.role == ROLE_SUPERVISOR and receiver.role == ROLE_USER:
        assigned = Application.query.filter_by(assigned_to_id=sender.id, applicant_id=receiver.id).first()
        return assigned is not None
    if sender.role == ROLE_USER and receiver.role == ROLE_SUPERVISOR:
        assigned = Application.query.filter_by(assigned_to_id=receiver.id, applicant_id=sender.id).first()
        return assigned is not None
    return False


def _get_allowed_recipients(user):
    """Return list of users that this user is allowed to message."""
    all_users = User.query.filter(User.id != user.id, User.is_active == True).all()
    return [u for u in all_users if _can_message(user, u)]


@messages_bp.route('/')
@login_required
def inbox():
    # Get all unique conversation partners
    sent_to = db.session.query(Message.receiver_id).filter_by(sender_id=current_user.id)
    recv_from = db.session.query(Message.sender_id).filter_by(receiver_id=current_user.id)
    partner_ids = set()
    for row in sent_to:
        partner_ids.add(row[0])
    for row in recv_from:
        partner_ids.add(row[0])

    conversations = []
    for pid in partner_ids:
        partner = db.session.get(User, pid)
        if not partner:
            continue
        # Last visible message in thread (respects soft-deletes)
        last_msg = (Message.query
                    .filter(
                        or_(
                            and_(Message.sender_id == current_user.id, Message.receiver_id == pid,
                                 Message.deleted_by_sender == False),
                            and_(Message.sender_id == pid, Message.receiver_id == current_user.id,
                                 Message.deleted_by_receiver == False)
                        )
                    )
                    .order_by(Message.created_at.desc())
                    .first())
        # Skip thread entirely if every message has been deleted by this user
        if not last_msg:
            continue
        unread = (Message.query
                  .filter_by(sender_id=pid, receiver_id=current_user.id, is_read=False)
                  .filter(Message.deleted_by_receiver == False)
                  .count())
        conversations.append({
            'partner': partner,
            'last_msg': last_msg,
            'unread': unread,
        })

    # Sort by last message date
    conversations.sort(key=lambda c: c['last_msg'].created_at if c['last_msg'] else 0, reverse=True)

    total_unread = sum(c['unread'] for c in conversations)
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    kpi_conversations  = len(conversations)
    kpi_unread         = total_unread
    kpi_today          = Message.query.filter(
        or_(Message.sender_id == current_user.id, Message.receiver_id == current_user.id),
        Message.created_at >= today_start
    ).count()
    kpi_total_msgs     = Message.query.filter(
        or_(Message.sender_id == current_user.id, Message.receiver_id == current_user.id)
    ).count()

    return render_template('messages/inbox.html',
                           conversations=conversations,
                           total_unread=total_unread,
                           kpi_conversations=kpi_conversations,
                           kpi_unread=kpi_unread,
                           kpi_today=kpi_today,
                           kpi_total_msgs=kpi_total_msgs)


@messages_bp.route('/thread/<int:partner_id>')
@login_required
def thread(partner_id):
    partner = db.get_or_404(User, partner_id)
    if not _can_message(current_user, partner) and not _can_message(partner, current_user):
        flash('You are not allowed to message this user.', 'danger')
        return redirect(url_for('messages.inbox'))

    messages_list = (Message.query
                     .filter(
                         or_(
                             and_(Message.sender_id == current_user.id, Message.receiver_id == partner_id,
                                  Message.deleted_by_sender == False),
                             and_(Message.sender_id == partner_id, Message.receiver_id == current_user.id,
                                  Message.deleted_by_receiver == False)
                         )
                     )
                     .order_by(Message.created_at.asc())
                     .all())

    # Mark all received as read
    (Message.query
     .filter_by(sender_id=partner_id, receiver_id=current_user.id, is_read=False)
     .update({'is_read': True}))
    db.session.commit()

    return render_template('messages/thread.html',
                           partner=partner,
                           messages_list=messages_list,
                           can_reply=_can_message(current_user, partner))


@messages_bp.route('/send', methods=['POST'])
@login_required
def send():
    receiver_id = request.form.get('receiver_id', type=int)
    body = request.form.get('body', '').strip()
    subject = request.form.get('subject', '').strip()
    next_url = request.form.get('next', '').strip()

    if not receiver_id or not body:
        flash('Message body is required.', 'danger')
        return redirect(next_url or url_for('messages.inbox'))

    receiver = db.get_or_404(User, receiver_id)
    if not _can_message(current_user, receiver):
        flash('You are not allowed to message this user.', 'danger')
        return redirect(next_url or url_for('messages.inbox'))

    msg = Message(
        sender_id=current_user.id,
        receiver_id=receiver_id,
        body=body,
        subject=subject or None,
    )
    db.session.add(msg)
    push_notification(
        receiver_id,
        f'New message from {current_user.full_name}: {body[:60]}{"…" if len(body) > 60 else ""}',
        url_for('messages.thread', partner_id=current_user.id),
        icon='bi-chat-dots-fill'
    )
    db.session.commit()
    flash('Message sent.', 'success')

    # Honour explicit next URL (e.g. application detail page)
    if next_url:
        return redirect(next_url)
    return redirect(url_for('messages.thread', partner_id=receiver_id))


@messages_bp.route('/compose')
@login_required
def compose():
    preset_receiver = request.args.get('to', type=int)
    channel = (request.args.get('channel') or '').strip().lower()
    recipients = [u for u in _get_allowed_recipients(current_user) if _recipient_matches_channel(u, channel)]
    preset_user = None
    if preset_receiver:
        preset_user = db.session.get(User, preset_receiver)
        if preset_user and channel and not _recipient_matches_channel(preset_user, channel):
            preset_user = None
    return render_template('messages/compose.html',
                           recipients=recipients,
                           preset_user=preset_user,
                           channel=channel)


@messages_bp.route('/export')
@login_required
def export():
    fmt = request.args.get('fmt', 'csv')
    msgs = (Message.query
            .filter(or_(Message.sender_id == current_user.id,
                        Message.receiver_id == current_user.id))
            .order_by(Message.created_at.desc()).all())
    headers = ['ID', 'From', 'To', 'Subject', 'Body', 'Date', 'Read']
    data = [[
        m.id,
        m.sender.full_name if m.sender else '',
        m.receiver.full_name if m.receiver else '',
        m.subject or '',
        m.body,
        m.created_at.strftime('%Y-%m-%d %H:%M'),
        'Yes' if m.is_read else 'No'
    ] for m in msgs]
    if fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Messages'
            header_fill = PatternFill('solid', fgColor='0F1923')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for row in data:
                ws.append([str(v) if v is not None else '' for v in row])
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
            import io as _io
            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            from flask import send_file
            return send_file(buf,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True,
                             download_name='messages.xlsx')
        except ImportError:
            pass
    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(headers)
    for row in data:
        writer.writerow([str(v) if v is not None else '' for v in row])
    return Response(si.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=messages.csv'})


@messages_bp.route('/<int:msg_id>/delete', methods=['POST'])
@login_required
def delete_message(msg_id):
    msg = db.get_or_404(Message, msg_id)
    next_url = request.form.get('next', '').strip()
    partner_id = None
    if msg.sender_id == current_user.id:
        msg.deleted_by_sender = True
        partner_id = msg.receiver_id
    elif msg.receiver_id == current_user.id:
        msg.deleted_by_receiver = True
        partner_id = msg.sender_id
    else:
        abort(403)
    db.session.commit()
    if next_url:
        return redirect(next_url)
    return redirect(url_for('messages.thread', partner_id=partner_id))
